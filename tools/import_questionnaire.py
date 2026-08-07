"""Разовый импорт спецификации опросника из xlsx в YAML-черновик.

Извлечение приблизительное: в исходном файле шкалы лежат в отдельной колонке,
часть шкал вписана прямо в текст вопроса, а подгруппы размечены слиянием ячеек.
Инструмент вытаскивает всё, что может, и помечает сомнительное строкой
"ПРОВЕРИТЬ". Результат дорабатывается руками один раз, после чего YAML
становится источником правды, а этот скрипт больше не нужен.

Запуск:
    uv run python tools/import_questionnaire.py \\
        Большой_интегральный_опросник.xlsx knowledge/questionnaire.draft.yaml
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from healthcoach.knowledge.import_xlsx import (
    is_section_heading,
    slugify,
    split_inline_scale,
)

SHEET_QUESTIONS = "ОПРОСНИК"
SHEET_KEY = "РЕЗУЛЬТАТ КЛЮЧ"


def main(xlsx: Path, out: Path) -> None:
    workbook = load_workbook(xlsx, data_only=True)
    sheet = workbook[SHEET_QUESTIONS]

    lines: list[str] = ['version: "1.0"', "blocks:"]
    current: str | None = None

    for row in sheet.iter_rows(min_row=1, max_col=8):
        a = row[0].value
        b = row[1].value
        h = row[7].value

        heading = is_section_heading(a) if isinstance(a, str) else None
        if heading and not b:
            title = heading.capitalize()
            current = slugify(title)
            lines += [
                f"  - id: {current}",
                f"    title: {title}",
                "    part: ПРОВЕРИТЬ  # организационная | клиническая | дополнительная",
                "    core: ПРОВЕРИТЬ  # true для ядра, false для дополнительных",
                "    scale: []  # ПРОВЕРИТЬ: перенести шкалу из колонки H",
                "    questions:",
            ]
            if isinstance(h, str) and h.strip():
                lines.append(f"    # шкала из колонки H: {h.strip()!r}")
            continue

        if current and isinstance(a, int) and isinstance(b, str):
            text, inline = split_inline_scale(b)
            lines.append(f"      - id: {current}.{a}")
            lines.append(f"        number: {a}")
            lines.append(f"        text: {text!r}")
            if inline:
                lines.append("        scale:")
                for option in inline:
                    lines.append(
                        f"          - {{score: {option['score']}, "
                        f"label: {option['label']!r}}}"
                    )

    lines += [
        "",
        "# ПРОВЕРИТЬ: подгруппы и пороги перенести с листа "
        f"{SHEET_KEY!r} вручную,",
        "# затем прогнать validate_questionnaire — он поймает пересечения и разрывы.",
    ]

    out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"черновик записан: {out}")
    print("дальше: заполнить пометки ПРОВЕРИТЬ и перенести пороги с листа ключа")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit(__doc__)
    main(Path(sys.argv[1]), Path(sys.argv[2]))
