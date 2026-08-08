"""Собрать knowledge/questionnaire.yaml из исходного xlsx.

Извлечение детерминированное: структура опросника размечена в самой таблице —
баннеры частей, заголовки блоков, подзаголовки подгрупп, строки «Сумма всех
баллов» и шкалы в колонке H. Скрипт читает эту разметку, а не переписывает
данные руками.

Судейская часть — соответствие блоков частям и признак ядра — задана таблицей
PARTS ниже: в самом файле этой информации нет, баннеры лишь разделяют секции.

Запуск:
    uv run python tools/build_questionnaire.py

Источник правды после сборки — knowledge/questionnaire.yaml. Скрипт нужен,
чтобы происхождение данных оставалось прослеживаемым.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from healthcoach.knowledge.degrees import degree_rank, degree_severity
from healthcoach.knowledge.import_xlsx import slugify, split_inline_scale
from healthcoach.knowledge.validation import parse_threshold_range

SHEET_QUESTIONS = "ОПРОСНИК"
SHEET_KEY = "РЕЗУЛЬТАТ КЛЮЧ"
SHEET_DASS = "DASS (показатели)"

BANNERS = {
    "ОРГАНИЗАЦИОННАЯ ЧАСТЬ": "организационная",
    "КЛИНИЧЕСКАЯ ЧАСТЬ": "клиническая",
    "ДОПОЛНИТЕЛЬНЫЕ ОПРОСНИКИ": "дополнительная",
}

CORE_PARTS = {"организационная", "клиническая"}

_LEADING_NUMBER = re.compile(r"^\s*\d+\.\s*")
_SUBSCALE = re.compile(r"^([А-Д])\.\s*(.+)$", re.DOTALL)
_SCALE_OPTION = re.compile(r"^\s*(\d+)\s*[-–—]\s*(.+)$")
_LOOKS_LIKE_RANGE = re.compile(r"^\s*(?:[<>]\s*\d+|\d+\s*[-–]\s*\d+)\s*$")

DASS_SUBSCALES = (
    ("депрессия", "Депрессия", "E"),
    ("тревожность", "Тревожность", "F"),
    ("стресс", "Стресс", "G"),
)

KEY_ALIASES = {
    # В исходной таблице заголовки блоков и строки ключа местами расходятся:
    # опечатки («ЭССЕНЦИАЛЬНЕ», «ИМУННАЯ», «опроник») и разные формулировки.
    "essencialne_nutrienty": "Эссенциальные нутриенты",
    "imunnaa_sistema": "Иммунная система",
    "korotkiy_oprosnik_po_ocenke_intoksikacii": "Короткий опроник по оценке интоксикации",
    "dass_oprosnik_depressia_trevoznost_stress": "DASS - депрессия, тревожность, стресс",
}

CANDIDA_TOTAL = "всего"
"""Пороги Candida относятся к сумме по всему опроснику, а не к секциям.

На листе ключа одни и те же значения продублированы в трёх строках подряд,
но секция «Прочие симптомы» набирает максимум 96 баллов и порог 141 для неё
недостижим в принципе. Максимум по всем трём секциям — 548, что сходится
с классическим опросником Крука. Подтверждено коучем: считаем сумму.
"""

CANDIDA_TOP_FIX = {"м": 141, "ж": 181}
"""Верхняя степень Candida записана в ключе как '<140' и '<180'.

Так она перекрывает низкую и среднюю и оставляет высокие баллы без степени.
По смыслу это «выше 140» и «выше 180»; подтверждено коучем. Средняя степень
заканчивается на 140 (180), поэтому высокая начинается со следующего балла.
"""

DASS_THRESHOLDS_DEFERRED = True
"""Пороги DASS не выставляются: они относятся к подшкалам, а не к сумме.

На листе ключа три строки — «Депрессия», «Тревожность», «Стресс», каждая
для своей 14-вопросной подшкалы. Пока подшкалы не разделены, любая из них,
применённая к сумме по 42 утверждениям, даёт заведомо неверную степень.
Разделение отложено коучем: разметка на скрытом листе относит каждый вопрос
сразу к двум подшкалам, а у «Стресса» тяжёлая степень повторяет умеренную.
"""

DASS_NOTE = """    # DASS: пороги не выставлены намеренно.
    # На листе ключа три строки порогов — «Депрессия», «Тревожность»,
    # «Стресс», — и каждая относится к своей подшкале из 14 утверждений,
    # а не к сумме по всем 42. Применённая к сумме, любая из них даёт
    # заведомо неверную степень, поэтому блок отдаёт сумму без степени.
    #
    # Чтобы включить разделение: замените подгруппу «весь» на три подгруппы
    # со списками вопросов и порогами ниже. Правится здесь, без изменений
    # в коде. Перед этим нужны данные от коуча: на скрытом листе
    # «DASS (показатели)» каждый вопрос отнесён сразу к двум подшкалам
    # (формулы вида =B{r}+D{r}), а у «Стресса» тяжёлая степень повторяет
    # умеренную (19-25), из-за чего баллы 26-33 не попадают никуда."""

DASS_DEGREES = (
    ("нормальный", 3),
    ("средний", 5),
    ("умеренный", 7),
    ("тяжелый", 9),
    ("очень тяжелый", 10),
)
"""У DASS на листе ключа пять градаций в колонках C, E, G, I, J."""


def _clean_title(raw: str) -> str:
    """Убрать порядковый номер и пояснение в скобках из заголовка блока."""
    title = _LEADING_NUMBER.sub("", raw.strip().splitlines()[0]).strip()
    return re.sub(r"\s{2,}", " ", title)


_CONNECTORS = {"и", "с", "в", "по", "на", "для", "или"}


def _is_block_heading(text: str, has_question_text: bool) -> bool:
    """Заголовок блока набран капсом; союзы вроде «и» могут быть строчными."""
    if has_question_text or not text:
        return False
    head = text.strip().splitlines()[0].strip()
    if head in BANNERS or head.startswith("("):
        return False
    words = [w for w in re.split(r"[\s/]+", _LEADING_NUMBER.sub("", head)) if w]
    significant = [w for w in words if w.lower() not in _CONNECTORS]
    if not significant:
        return False
    letters = [c for w in significant for c in w if c.isalpha()]
    return len(letters) > 4 and all(c.isupper() for c in letters)


def _collect_scale(sheet, start_row: int, end_row: int) -> list[dict]:
    """Собрать шкалу из колонки H в пределах диапазона строк.

    Шкала относится к подгруппе, а не к блоку: у «Питания» части А и Б
    заданы противоположно (0 — не употребляю против 0 — ежедневно),
    у Candida каждая секция считается по-своему (35, 3/6/9, 1/2/3),
    у QEESI индекс маскировки отличается от остальных секций.
    """
    options: list[dict] = []
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row, 8).value
        if not isinstance(cell, str):
            continue
        for line in cell.splitlines():
            match = _SCALE_OPTION.match(line)
            if match:
                options.append(
                    {"score": int(match.group(1)), "label": match.group(2).strip()}
                )
    seen: set[int] = set()
    unique: list[dict] = []
    for option in options:
        if option["score"] not in seen:
            seen.add(option["score"])
            unique.append(option)
    return sorted(unique, key=lambda o: o["score"])


def _read_questions(sheet) -> tuple[list[dict], dict[str, list[dict]]]:
    """Пройти лист опросника и разложить вопросы по блокам и подгруппам."""
    blocks: list[dict] = []
    part: str | None = None
    block: dict | None = None
    subscale: dict | None = None

    for row in range(1, sheet.max_row + 1):
        a = sheet.cell(row, 1).value
        b = sheet.cell(row, 2).value
        text = a.strip() if isinstance(a, str) else ""
        head = text.splitlines()[0].strip() if text else ""

        if head in BANNERS:
            part = BANNERS[head]
            continue

        if isinstance(a, int) and isinstance(b, str) and block is not None:
            if subscale is None:
                subscale = {"id": "весь", "title": "Весь блок", "questions": []}
                block["subscales"].append(subscale)
            question_text, inline_scale = split_inline_scale(b)
            subscale["questions"].append(
                {
                    "number": a,
                    "text": question_text,
                    "scale": inline_scale,
                    "row": row,
                }
            )
            continue

        if text.startswith("Сумма"):
            continue

        if _is_block_heading(text, bool(b)):
            title = _clean_title(text)
            block = {
                "id": slugify(title),
                "title": title,
                "part": part,
                "core": part in CORE_PARTS,
                "row": row,
                "subscales": [],
            }
            blocks.append(block)
            subscale = None
            continue

        match = _SUBSCALE.match(text) if text else None
        if match and block is not None:
            letter = match.group(1)
            label = _clean_title(match.group(2).split(" - ")[0])
            subscale = {
                "id": letter.lower(),
                "title": f"{letter}. {label}",
                "questions": [],
                "row": row,
            }
            block["subscales"].append(subscale)

    return blocks, {}


def _read_dass_membership(sheet) -> dict[str, list[int]]:
    """Прочитать принадлежность вопросов DASS к трём подшкалам.

    На листе номера вопросов идут двумя парами колонок (A/B и C/D), а колонки
    E, F, G помечают, к какой подшкале относится вопрос из этой строки.
    """
    membership: dict[str, list[int]] = {key: [] for key, _, _ in DASS_SUBSCALES}
    columns = {"E": 5, "F": 6, "G": 7}

    for row in range(2, sheet.max_row + 1):
        numbers = [
            sheet.cell(row, col).value
            for col in (1, 3)
            if isinstance(sheet.cell(row, col).value, int)
        ]
        if not numbers:
            continue
        for key, _, column in DASS_SUBSCALES:
            if sheet.cell(row, columns[column]).value is not None:
                membership[key].extend(numbers)

    return {key: sorted(set(values)) for key, values in membership.items()}


def _read_thresholds(sheet) -> dict[str, list[dict]]:
    """Прочитать пороги с листа ключа.

    Колонки: E/F — низкая, G/H — средняя, I/J — высокая. Вторая колонка пары
    используется только там, где пороги различаются по полу.
    """
    degrees = (("низкая", 5, 6), ("средняя", 7, 8), ("высокая", 9, 10))
    result: dict[str, list[dict]] = {}
    current: str | None = None
    in_dass = False

    for row in range(1, sheet.max_row + 1):
        a = sheet.cell(row, 1).value
        b = sheet.cell(row, 2).value

        if isinstance(a, str) and a.strip() and a.strip() not in BANNERS:
            head = _clean_title(a)
            if head and not head.startswith("КЛЮЧ"):
                current = head
                in_dass = head.upper().startswith("DASS")

        if current is None:
            continue

        thresholds: list[dict] = []

        if in_dass:
            for degree, column in DASS_DEGREES:
                raw = sheet.cell(row, column).value
                if isinstance(raw, str) and _LOOKS_LIKE_RANGE.match(raw):
                    low, high = parse_threshold_range(raw)
                    thresholds.append(
                        {"degree": degree, "min": low, "max": high, "sex": None}
                    )
            if thresholds and isinstance(b, str) and b.strip():
                result.setdefault(f"{current}|{_clean_title(b)}", []).extend(thresholds)
            continue

        for degree, male_col, female_col in degrees:
            male_raw = sheet.cell(row, male_col).value
            female_raw = sheet.cell(row, female_col).value
            if not isinstance(male_raw, str) or not _LOOKS_LIKE_RANGE.match(male_raw):
                continue
            if isinstance(female_raw, str) and _LOOKS_LIKE_RANGE.match(female_raw):
                for raw, sex in ((male_raw, "м"), (female_raw, "ж")):
                    low, high = parse_threshold_range(raw)
                    thresholds.append(
                        {"degree": degree, "min": low, "max": high, "sex": sex}
                    )
            else:
                low, high = parse_threshold_range(male_raw)
                thresholds.append(
                    {"degree": degree, "min": low, "max": high, "sex": None}
                )

        if not thresholds:
            continue

        key = current
        if isinstance(b, str) and b.strip():
            key = f"{current}|{_clean_title(b)}"
        result.setdefault(key, []).extend(thresholds)

    return result


def _quote(text: str) -> str:
    escaped = str(text).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def _match_thresholds(
    block: dict, subscale: dict, thresholds: dict[str, list[dict]]
) -> list[dict]:
    """Найти пороги подгруппы на листе ключа по названию блока и подгруппы."""
    def key_slug(key: str) -> str:
        head = key.split("|")[0]
        head = re.sub(r"^\s*Блок\s+", "", head, flags=re.IGNORECASE)
        return slugify(head)

    alias = KEY_ALIASES.get(block["id"])
    want = slugify(alias) if alias else slugify(
        re.sub(r"^\s*Блок\s+", "", block["title"], flags=re.IGNORECASE)
    )
    candidates = [key for key in thresholds if key_slug(key) == want]
    if not candidates:
        candidates = [
            key
            for key in thresholds
            if key_slug(key)[:14] == want[:14] and len(want) > 6
        ]
    if not candidates:
        return []

    if len(block["subscales"]) == 1 or subscale["id"] == CANDIDA_TOTAL:
        # У Candida одни и те же пороги продублированы в трёх строках ключа —
        # берём первую, она относится к сумме по всему опроснику.
        return thresholds[sorted(candidates)[0]]

    letter = subscale["id"].upper()
    for key in candidates:
        tail = key.split("|")[1] if "|" in key else ""
        if tail.strip().startswith(f"{letter}."):
            return thresholds[key]
    return []


def _emit(blocks: list[dict], thresholds: dict[str, list[dict]]) -> str:
    lines = ['version: "1.0"', "blocks:"]
    for block in blocks:
        lines += [
            f"  - id: {block['id']}",
            f"    title: {_quote(block['title'])}",
            f"    part: {block['part']}",
            f"    core: {str(block['core']).lower()}",
            "    scale:",
        ]
        for option in block["scale"]:
            lines.append(
                f"      - {{score: {option['score']}, label: {_quote(option['label'])}}}"
            )
        lines.append("    questions:")
        for subscale in block["subscales"]:
            for question in subscale["questions"]:
                qid = f"{block['id']}.{question['number']}"
                if len(block["subscales"]) > 1:
                    qid = f"{block['id']}.{subscale['id']}.{question['number']}"
                lines += [
                    f"      - id: {qid}",
                    f"        number: {question['number']}",
                    f"        text: {_quote(question['text'])}",
                ]
                if question["scale"]:
                    lines.append("        scale:")
                    for option in question["scale"]:
                        lines.append(
                            f"          - {{score: {option['score']}, "
                            f"label: {_quote(option['label'])}}}"
                        )
        lines.append("    subscales:")
        if block["id"] == "oprosnik_candida":
            lines.append(
                "    # Секции считаются отдельно, но степень выносится по общей"
                "\n    # сумме: порог 141/181 недостижим для секции «Прочие симптомы»"
                "\n    # (её максимум 96). Максимум по всем секциям — 548."
            )
        if block["id"].startswith("dass"):
            lines.append(DASS_NOTE)
            for key in sorted(thresholds):
                if key.upper().startswith("DASS"):
                    name = key.split("|")[1]
                    formatted = "; ".join(
                        f"{t['degree']} {t['min']}..{'' if t['max'] is None else t['max']}"
                        for t in thresholds[key]
                    )
                    lines.append(f"    # {name}: {formatted}")
        emitted = list(block["subscales"])
        if block["id"] == "oprosnik_candida":
            emitted.append(
                {
                    "id": CANDIDA_TOTAL,
                    "title": "Всего по опроснику",
                    "questions": [q for s in block["subscales"] for q in s["questions"]],
                    "sections": [s["id"] for s in block["subscales"]],
                }
            )

        for subscale in emitted:
            qids = []
            for question in subscale["questions"]:
                section = subscale.get("sections") and next(
                    s["id"]
                    for s in block["subscales"]
                    if question in s["questions"]
                )
                owner = section or subscale["id"]
                qid = f"{block['id']}.{question['number']}"
                if len(block["subscales"]) > 1:
                    qid = f"{block['id']}.{owner}.{question['number']}"
                qids.append(qid)
            lines += [
                f"      - id: {subscale['id']}",
                f"        title: {_quote(subscale['title'])}",
                "        question_ids:",
            ]
            lines += [f"          - {qid}" for qid in qids]
            if block["id"] == "oprosnik_candida" and subscale["id"] != CANDIDA_TOTAL:
                # Пороги вынесены в подгруппу «всего»; секции дают только сумму.
                lines.append("        thresholds: []")
                continue
            if DASS_THRESHOLDS_DEFERRED and block["id"].startswith("dass"):
                # Пороги на листе ключа — это пороги подшкал, а не суммы по
                # 42 утверждениям; см. DASS_NOTE и DASS_THRESHOLDS_DEFERRED.
                lines.append("        thresholds: []")
                continue
            found = _match_thresholds(block, subscale, thresholds)
            if "qeesi" in block["id"]:
                # Верхние границы 100 и 10 — потолок шкалы, а не порог: они
                # в точности равны максимально достижимой сумме секции
                # (10 вопросов по 10 баллов и 10 вопросов по 1 баллу).
                # Открываем степень сверху — на достижимых баллах это ничего
                # не меняет, но снимает мнимую находку валидатора.
                found = [
                    {**t, "max": None} if t["degree"] == "высокая" else t
                    for t in found
                ]
            if block["id"] == "oprosnik_candida":
                found = [
                    {**t, "min": CANDIDA_TOP_FIX[t["sex"]], "max": None}
                    if t["degree"] == "высокая" and t["sex"] in CANDIDA_TOP_FIX
                    else t
                    for t in found
                ]
            if not found:
                lines.append("        thresholds: []")
                continue
            lines.append("        thresholds:")
            for threshold in found:
                low = "null" if threshold["min"] is None else threshold["min"]
                high = "null" if threshold["max"] is None else threshold["max"]
                sex = "null" if threshold["sex"] is None else threshold["sex"]
                lines.append(
                    f"          - {{degree: {threshold['degree']}, min: {low}, "
                    f"max: {high}, sex: {sex}}}"
                )
    return "\n".join(lines) + "\n"


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    workbook = load_workbook(root / "Большой_интегральный_опросник.xlsx", data_only=True)

    blocks, _ = _read_questions(workbook[SHEET_QUESTIONS])
    sheet = workbook[SHEET_QUESTIONS]
    unscaled: list[str] = []

    for block in blocks:
        # Шкалу ищем по каждой подгруппе отдельно: в разных секциях одного
        # блока она может быть разной и даже противоположной.
        for index, subscale in enumerate(block["subscales"]):
            rows = [q["row"] for q in subscale["questions"]]
            if not rows:
                subscale["scale"] = []
                continue
            start = subscale.get("row", block["row"] if index == 0 else min(rows))
            subscale["scale"] = _collect_scale(sheet, start, max(rows))

        with_scale = [s["scale"] for s in block["subscales"] if s["scale"]]
        block["scale"] = with_scale[0] if with_scale else []

        if not block["scale"]:
            # Шкала не выписана в колонке H, но может быть вписана в каждый
            # вопрос — тогда собираем её объединением.
            inline: dict[int, str] = {}
            questions = [q for sub in block["subscales"] for q in sub["questions"]]
            if questions and all(q["scale"] for q in questions):
                for question in questions:
                    for option in question["scale"]:
                        inline.setdefault(option["score"], option["label"])
                block["scale"] = [
                    {"score": score, "label": inline[score]} for score in sorted(inline)
                ]

        # Подгруппа со своей шкалой прописывает её каждому своему вопросу,
        # иначе вопрос унаследовал бы шкалу соседней секции.
        for subscale in block["subscales"]:
            if not subscale["scale"] or subscale["scale"] == block["scale"]:
                continue
            for question in subscale["questions"]:
                if not question["scale"]:
                    question["scale"] = list(subscale["scale"])

        # Подгруппа без собственной шкалы (H-колонка пуста для неё) молча
        # получает шкалу блока, унаследованную от другой подгруппы — это тот
        # самый механизм, что подставил Candida-секции А чужую шкалу 3/6/9.
        # Безопасно только пока у каждого вопроса подгруппы есть инлайн-шкала.
        for subscale in block["subscales"]:
            if subscale["scale"]:
                continue
            missing = [q["number"] for q in subscale["questions"] if not q["scale"]]
            if not missing:
                continue
            unscaled.append(
                f"{block['id']}/{subscale['id']}: своей шкалы нет, а у вопросов "
                f"{missing} нет инлайн-шкалы — подгруппа молча унаследует шкалу "
                f"другой подгруппы блока"
            )

        for subscale in block["subscales"]:
            for question in subscale["questions"]:
                if not question["scale"] and not block["scale"]:
                    unscaled.append(f"{block['id']}/{subscale['id']}.{question['number']}")

    if unscaled:
        print("НЕ НАЙДЕНА ШКАЛА:", unscaled[:10], f"(всего {len(unscaled)})")
        return 1

    thresholds = _read_thresholds(workbook[SHEET_KEY])

    # Названия степеней здесь — независимая копия словаря healthcoach.knowledge
    # .degrees; имя, которого нет в DEGREE_ORDER/DEGREE_SEVERITY, тихо получит
    # неизвестную тяжесть при сортировке находок. Проверяем перед записью.
    degree_names = {name for name, _ in DASS_DEGREES}
    for rows in thresholds.values():
        degree_names.update(t["degree"] for t in rows)
    unknown_degrees = sorted(
        name
        for name in degree_names
        if degree_rank(name) is None or degree_severity(name) is None
    )
    if unknown_degrees:
        print("НЕИЗВЕСТНАЯ СТЕПЕНЬ:", unknown_degrees)
        return 1

    out = root / "knowledge" / "questionnaire.yaml"
    out.write_text(_emit(blocks, thresholds), encoding="utf-8")

    total = sum(len(s["questions"]) for b in blocks for s in b["subscales"])
    missing_scale = [b["id"] for b in blocks if not b["scale"]]
    missing_thresholds = [
        f"{b['id']}/{s['id']}"
        for b in blocks
        for s in b["subscales"]
        if (DASS_THRESHOLDS_DEFERRED and b["id"].startswith("dass"))
        or not _match_thresholds(b, s, thresholds)
    ]

    print(f"записано: {out}")
    print(f"блоков: {len(blocks)}, вопросов: {total}")
    print(f"без шкалы: {missing_scale or 'нет'}")
    print(f"без порогов: {missing_thresholds or 'нет'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
